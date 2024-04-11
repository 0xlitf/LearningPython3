from openpyxl import load_workbook
import time


def time_decorator(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        print(f"{func.__name__} executed in {end_time - start_time:.2f} seconds")
        return result
    return wrapper


@time_decorator
def read_excel():
    # 加载工作簿
    workbook = load_workbook(filename='23122513_20231225153028.xlsx')

    # 通过名称选择工作表
    sheet = workbook['Sheet1']

    # 或者选择当前活动的工作表
    # sheet = workbook.active

    # 读取A1单元格的值
    # cell_value = sheet['A1'].value
    # print(cell_value)

    # 或者通过行和列的索引读取（注意：行和列的索引从1开始）
    # cell_value = sheet.cell(row=1, column=1).value
    # print(cell_value)

    # 遍历第一行的所有单元格
    # for cell in sheet[1]:
    #     print(cell.value)

    # 遍历A列的所有单元格
    # for cell in sheet['A']:
    #     print(cell.value)

    # 遍历所有行
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)

    return rows
    # 遍历所有列
    # for column in sheet.iter_cols(values_only=True):
    #     print(column)


if __name__ == '__main__':
    rows = read_excel()
    print(len(rows))
